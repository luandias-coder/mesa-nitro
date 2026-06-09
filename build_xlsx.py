#!/usr/bin/env python3
"""Gera planilha .xlsx das lives parseadas/resolvidas para validacao 1x1.
Le signals_resolved_<data>.json e monta:
  - aba "Operacoes do dia" (entrada/saida/ajuste)
  - aba "Posicoes em aberto" (snapshot citado na live)
Cada linha tem link direto pro segundo exato do video.
"""
import json
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

# data -> video id do YouTube
VIDEOS = {
    "2026-06-02": "CTDht2slsvs",
    "2026-06-03": "T0eVXLngc1k",
    "2026-06-05": "3mXB4MVn004",
    "2026-06-08": "Nzk00uLlR-s",
}

def ts_to_seconds(ts):
    if not ts:
        return None
    parts = [int(p) for p in str(ts).split(":")]
    if len(parts) == 3:
        h, m, s = parts
    elif len(parts) == 2:
        h, m, s = 0, parts[0], parts[1]
    else:
        return None
    return h * 3600 + m * 60 + s

def link_for(data, ts):
    vid = VIDEOS.get(data)
    secs = ts_to_seconds(ts)
    if not vid or secs is None:
        return None
    return f"https://www.youtube.com/watch?v={vid}&t={secs}s"

# --- estilos ---
HDR_FILL = PatternFill("solid", fgColor="1F2A44")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TICKER_FILL = PatternFill("solid", fgColor="FFF2CC")
TICKER_FONT = Font(bold=True, color="7F6000")
LINK_FONT = Font(color="1155CC", underline="single")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ACAO_LABEL = {
    "entrada": "ENTRADA",
    "saida": "SAIDA",
    "ajuste_delta": "AJUSTE DELTA",
}

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def write_sheet(ws, headers, rows, ticker_col, link_col, conf_col, widths):
    ws.append(headers)
    style_header(ws, len(headers))
    for r in rows:
        ws.append(r["values"])
        ridx = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=ridx, column=c).border = BORDER
            ws.cell(row=ridx, column=c).alignment = LEFT
        # link
        if r.get("link"):
            lcell = ws.cell(row=ridx, column=link_col)
            lcell.hyperlink = r["link"]
            lcell.font = LINK_FONT
            lcell.alignment = CENTER
        # ticker real destacado
        tcell = ws.cell(row=ridx, column=ticker_col)
        tcell.fill = TICKER_FILL
        tcell.font = TICKER_FONT
        tcell.alignment = CENTER
        # destaque confianca != alta
        if r.get("conf") and r["conf"] != "alta":
            for c in range(1, len(headers) + 1):
                ws.cell(row=ridx, column=c).fill = WARN_FILL
            tcell.fill = TICKER_FILL  # mantem ticker amarelo
    # larguras
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def load_all():
    files = sorted(glob.glob(os.path.join(HERE, "signals_resolved_*.json")))
    return [json.load(open(f)) for f in files]

def main():
    docs = load_all()
    wb = Workbook()

    # ===== ABA 1: OPERACOES DO DIA =====
    ws1 = wb.active
    ws1.title = "Operacoes do dia"
    headers1 = [
        "Data", "Hora", "Link video", "Ativo", "Acao", "Tipo",
        "Codigo falado", "TICKER REAL (home broker)", "Strike real",
        "Preco live (R$)", "Ultimo grade (R$)", "Delta", "Vencimento",
        "Metodo", "Confianca", "Trecho da fala", "Motivo", "OK? (validar)",
    ]
    rows1 = []
    for d in docs:
        data = d.get("data")
        for s in d.get("sinais", []):
            res = s.get("resolucao", {}) or {}
            link = link_for(data, s.get("timestamp"))
            rows1.append({
                "values": [
                    data,
                    s.get("timestamp", ""),
                    "▶ abrir" if link else "",
                    s.get("underlying", ""),
                    ACAO_LABEL.get(s.get("acao"), s.get("acao", "")),
                    s.get("tipo_opcao", ""),
                    s.get("codigo_falado", ""),
                    s.get("codigo_resolvido", res.get("codigo_resolvido", "")),
                    res.get("strike_real", s.get("strike", "")),
                    s.get("preco_sugerido", ""),
                    res.get("ultimo_hoje", ""),
                    s.get("delta", res.get("delta_hoje", "")),
                    res.get("vencimento", s.get("vencimento", "")),
                    res.get("metodo", ""),
                    res.get("confianca", s.get("confianca", "")),
                    s.get("trecho", ""),
                    s.get("motivo", ""),
                    "",
                ],
                "link": link,
                "conf": res.get("confianca", s.get("confianca")),
            })
    widths1 = [11, 7, 9, 8, 13, 6, 16, 24, 11, 12, 14, 7, 12, 13, 10, 46, 46, 12]
    write_sheet(ws1, headers1, rows1, ticker_col=8, link_col=3, conf_col=15, widths=widths1)

    # ===== ABA 2: POSICOES EM ABERTO =====
    ws2 = wb.create_sheet("Posicoes em aberto")
    headers2 = [
        "Data", "Hora", "Link video", "Ativo", "Tipo", "Codigo falado",
        "TICKER REAL (home broker)", "Strike real", "Preco live (R$)",
        "Ultimo grade (R$)", "Vencimento", "Metodo", "Confianca",
        "Nota", "OK? (validar)",
    ]
    rows2 = []
    for d in docs:
        data = d.get("data")
        for p in d.get("posicoes_em_aberto", []):
            res = p.get("resolucao", {}) or {}
            link = link_for(data, p.get("timestamp"))
            nota = res.get("nota") or p.get("nota", "")
            rows2.append({
                "values": [
                    data,
                    p.get("timestamp", ""),
                    "▶ abrir" if link else "",
                    p.get("underlying", ""),
                    p.get("tipo", ""),
                    p.get("codigo_falado", ""),
                    p.get("codigo_resolvido", res.get("codigo_resolvido", "")),
                    res.get("strike_real", ""),
                    p.get("preco", ""),
                    res.get("ultimo_hoje", ""),
                    res.get("vencimento", ""),
                    res.get("metodo", ""),
                    res.get("confianca", ""),
                    nota,
                    "",
                ],
                "link": link,
                "conf": res.get("confianca"),
            })
    widths2 = [11, 7, 9, 8, 6, 18, 24, 11, 12, 14, 12, 13, 10, 50, 12]
    write_sheet(ws2, headers2, rows2, ticker_col=7, link_col=3, conf_col=13, widths=widths2)

    # ===== ABA 3: AVISOS / AMBIGUIDADES =====
    ws3 = wb.create_sheet("Avisos")
    ws3.append(["Data", "Aviso / ambiguidade a confirmar"])
    style_header(ws3, 2)
    for d in docs:
        for a in d.get("avisos", []):
            ws3.append([d.get("data"), a])
            ridx = ws3.max_row
            for c in (1, 2):
                ws3.cell(row=ridx, column=c).border = BORDER
                ws3.cell(row=ridx, column=c).alignment = LEFT
    ws3.column_dimensions["A"].width = 11
    ws3.column_dimensions["B"].width = 110
    ws3.freeze_panes = "A2"

    out = os.path.join(HERE, "validacao_lives.xlsx")
    wb.save(out)
    print(f"OK -> {out}")
    print(f"Operacoes: {len(rows1)} | Posicoes: {len(rows2)}")

if __name__ == "__main__":
    main()
